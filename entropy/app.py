import os
import io
import json
import psycopg2
from dotenv import load_dotenv

load_dotenv()
import psycopg2.extras
import pandas as pd
import numpy as np
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

# Database Configuration — credentials loaded from environment only
DB_CONFIG = {
    "host": os.environ.get("PG_HOST"),
    "port": int(os.environ.get("PG_PORT", "5432")),
    "user": os.environ.get("PG_USER"),
    "password": os.environ.get("PG_PASSWORD"),
    "dbname": os.environ.get("PG_DATABASE"),
}

def get_db_connection():
    if not all([DB_CONFIG["host"], DB_CONFIG["user"], DB_CONFIG["password"], DB_CONFIG["dbname"]]):
        raise RuntimeError("Database environment variables (PG_HOST, PG_USER, PG_PASSWORD, PG_DATABASE) are not set.")
    return psycopg2.connect(**DB_CONFIG)

def calculate_shannon_entropy(series):
    """Calculate Shannon entropy for a time series of quantities/revenues (12 months)."""
    total = series.sum()
    if total <= 0:
        return 0.0
    p = series / total
    p = p[p > 0]  # Ignore zero-probability months
    entropy = -np.sum(p * np.log2(p))
    return float(entropy)

def run_entropy_analysis(fy, segment='All', min_months=3, min_qty=0, winsorize_pct=99, abc_a=0.80, abc_b=0.15, adi_cutoff=1.32, cv2_cutoff=0.49):
    """
    Run core SKU temporal entropy and demand pattern analysis.
    Performs data cleaning, winsorization, zero-filling, and classification.
    """
    conn = get_db_connection()
    
    query = r"""
    SELECT 
        "Module WP" AS sku,
        "Mat Desc" AS mat_desc,
        "Invoice date" AS invoice_date,
        "SalesQty" AS qty,
        COALESCE("Revenue", "Taxable Value") AS revenue_val,
        "Year" AS fy,
        "Month" AS month,
        "Plant" AS plant,
        "Segment" AS segment,
        "Cust_name" AS cust_name
    FROM revenue
    WHERE "SalesQty" > 0 AND "Module WP" IS NOT NULL AND "Year" = %s;
    """
    
    df = pd.read_sql_query(query, conn, params=(fy,))
    conn.close()
    
    if df.empty:
        return pd.DataFrame(), {}
        
    if segment != 'All':
        df = df[df['segment'] == segment]
        
    if df.empty:
        return pd.DataFrame(), {}
        
    # 1. Non-winsorized data for true demand distribution (Entropy, ADI)
    grouped_true = df.groupby(['sku', 'month']).agg({
        'qty': 'sum',
        'revenue_val': 'sum'
    }).reset_index()

    # 2. Winsorized data for volatility (CV, CV²) metrics of REVENUE
    df_volatility = df.copy()
    if winsorize_pct > 0 and winsorize_pct < 100:
        r_limit = np.percentile(df_volatility['revenue_val'], winsorize_pct)
        df_volatility['revenue_val'] = np.clip(df_volatility['revenue_val'], None, r_limit)

    grouped_vol = df_volatility.groupby(['sku', 'month']).agg({
        'revenue_val': 'sum'
    }).reset_index()

    # Complete 12-month skeleton
    months_fy = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    unique_skus = df['sku'].unique()
    
    sku_desc = df.groupby('sku')['mat_desc'].apply(lambda x: x.mode().iloc[0] if not x.mode().empty else 'Unknown').to_dict()
    
    index = pd.MultiIndex.from_product([unique_skus, months_fy], names=['sku', 'month'])
    skeleton = pd.DataFrame(index=index).reset_index()
    
    merged_true = pd.merge(skeleton, grouped_true, on=['sku', 'month'], how='left').fillna(0)
    merged_vol = pd.merge(skeleton, grouped_vol, on=['sku', 'month'], how='left').fillna(0)
    
    # Store winsorized revenue series for CV lookup
    vol_by_sku = {sku: g.set_index('month').loc[months_fy]['revenue_val'].values for sku, g in merged_vol.groupby('sku')}
    
    sku_metrics = []
    max_h = np.log2(12)
    
    for sku, group in merged_true.groupby('sku'):
        group = group.set_index('month').loc[months_fy].reset_index()
        
        q_series = group['qty'].values
        r_series = group['revenue_val'].values
        
        total_qty = np.sum(q_series)
        total_revenue = np.sum(r_series)
        
        # If no active revenue, skip
        if total_revenue <= 0:
            continue
            
        # Non-zero months are months with active revenue
        nonzero_months = np.sum(r_series > 0)
        
        if nonzero_months < min_months:
            continue
            
        # Entropy of Revenue (using True non-winsorized series)
        h_rev = calculate_shannon_entropy(pd.Series(r_series))
        h_norm_rev = h_rev / max_h
        
        # Volatility of Revenue (using Winsorized series)
        r_series_vol = vol_by_sku.get(sku, r_series)
        mean_rev = np.mean(r_series_vol)
        std_rev = np.std(r_series_vol, ddof=0)
        cv = std_rev / mean_rev if mean_rev > 0 else 0.0
        cv2 = cv ** 2
        
        # ADI of Revenue
        adi = 12.0 / nonzero_months if nonzero_months > 0 else 99.0
        
        # Entropy Level of Revenue
        if h_norm_rev < 0.30:
            ent_level = 'Low'
        elif h_norm_rev < 0.60:
            ent_level = 'Medium'
        elif h_norm_rev < 0.80:
            ent_level = 'High'
        else:
            ent_level = 'Very High'
            
        sku_metrics.append({
            'Module WP': str(sku),
            'Mat Desc': sku_desc.get(sku, 'Unknown SKU'),
            'FY Qty': float(total_qty),
            'FY Revenue': float(total_revenue),
            'Entropy': float(h_rev),
            'H_norm': float(h_norm_rev),
            'NonZeroMonths': int(nonzero_months),
            'Entropy_Level': ent_level
        })
        
    if not sku_metrics:
        return pd.DataFrame(), {}
        
    df_res = pd.DataFrame(sku_metrics)
    
    # ABC Pareto
    df_res = df_res.sort_values(by='FY Revenue', ascending=False)
    total_rev = df_res['FY Revenue'].sum()
    
    if total_rev > 0:
        df_res['cum_rev'] = df_res['FY Revenue'].cumsum()
        df_res['cum_rev_pct'] = df_res['cum_rev'] / total_rev
    else:
        df_res['cum_rev_pct'] = 0.0
        
    abc_classes = []
    running_pct = 0.0
    for idx, row in df_res.iterrows():
        prev_pct = running_pct
        running_pct = row['cum_rev_pct']
        
        if prev_pct <= abc_a:
            abc_class = 'A'
        elif prev_pct <= (abc_a + abc_b):
            abc_class = 'B'
        else:
            abc_class = 'C'
        abc_classes.append(abc_class)
        
    df_res['ABC Class'] = abc_classes
    df_res = df_res.drop(columns=['cum_rev', 'cum_rev_pct'])
    
    # Combined segment and Recommendations
    overall_segments = []
    recs = []
    
    for idx, row in df_res.iterrows():
        abc = row['ABC Class']
        ent = row['Entropy_Level']
        
        overall = f"{abc}-{ent}"
        overall_segments.append(overall)
        
        # Recommendations
        if abc == 'A':
            if ent == 'Low':
                rec = "Strategic Stable Core SKU. High revenue, highly predictable. Maintain lean buffer stock; optimize supply chains."
            elif ent == 'Medium':
                rec = "Core Seasonal SKU. High revenue with moderate/predictable seasonality. Schedule production/sales runs to match seasonal peaks."
            else: # High / Very High
                rec = "Critical Volatile SKU. High revenue but highly erratic/unstable. Keep strategic safety stock or operate on Make-to-Order (MTO)."
        elif abc == 'B':
            if ent == 'Low':
                rec = "Mid-Tier Stable SKU. Regular, predictable revenue. Set standard automated replenishment rules with moderate safety levels."
            elif ent == 'Medium':
                rec = "Mid-Tier Seasonal SKU. Medium revenue with distinct seasons. Maintain normal safety margins; align buffers before seasonal peaks."
            else: # High / Very High
                rec = "Risky Mid-Tier SKU. Erratic revenue over time. Review safety stock frequently; consider rationalization or MTO."
        else: # C
            if ent == 'Low':
                rec = "Niche Stable SKU. Small revenue contribution but highly predictable. Keep minimal stock or operate on standard visual kanban."
            else: # Medium / High / Very High
                rec = "Erratic Slow-Moving SKU. Minimal revenue contribution and highly unstable. Transition strictly to Make-to-Order (MTO) or review for rationalization."
                
        recs.append(rec)
        
    df_res['Overall_Segment'] = overall_segments
    df_res['Actionable Recommendation'] = recs
    
    # Calculate summary KPIs
    kpis = {
        "total_skus": len(df_res),
        "total_revenue": float(total_rev),
        "total_qty": float(df_res['FY Qty'].sum()),
        "avg_entropy": float(df_res['H_norm'].mean()) if len(df_res) > 0 else 0,
        "abc_counts": df_res['ABC Class'].value_counts().to_dict(),
        "abc_revenue": df_res.groupby('ABC Class')['FY Revenue'].sum().to_dict()
    }
    
    # Ensure keys exist
    for key in ['A', 'B', 'C']:
        kpis['abc_counts'].setdefault(key, 0)
        kpis['abc_revenue'].setdefault(key, 0.0)
        
    return df_res, kpis

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/filters')
def get_filters():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # FY Years
        cursor.execute('SELECT DISTINCT "Year" FROM revenue WHERE "Year" IS NOT NULL ORDER BY "Year" DESC;')
        years = [r[0] for r in cursor.fetchall()]
        
        # Segments
        cursor.execute('SELECT DISTINCT "Segment" FROM revenue WHERE "Segment" IS NOT NULL ORDER BY "Segment";')
        segments = [r[0] for r in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "status": "success",
            "years": years,
            "plants": [],
            "segments": segments,
            "customers": []
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Database error: {str(e)}"
        }), 500

@app.route('/api/analyze')
def analyze():
    try:
        fy = request.args.get('fy', '2025-26')
        segment = request.args.get('segment', 'All')
        min_months = int(request.args.get('min_months', 3))
        min_qty = float(request.args.get('min_qty', 0))
        winsorize_pct = int(request.args.get('winsorize_pct', 99))
        abc_a = float(request.args.get('abc_a', 0.80))
        abc_b = float(request.args.get('abc_b', 0.15))
        adi_cutoff = float(request.args.get('adi_cutoff', 1.32))
        cv2_cutoff = float(request.args.get('cv2_cutoff', 0.49))
        
        df, kpis = run_entropy_analysis(
            fy=fy, segment=segment,
            min_months=min_months, min_qty=min_qty, 
            winsorize_pct=winsorize_pct, abc_a=abc_a, abc_b=abc_b,
            adi_cutoff=adi_cutoff, cv2_cutoff=cv2_cutoff
        )
        
        if df.empty:
            return jsonify({
                "status": "success",
                "skus": [],
                "kpis": {},
                "heatmap": {"x": [], "y": [], "z": []}
            })
            
        skus_data = df.to_dict(orient='records')
        
        # Calculate ABC vs Entropy Heatmap (Revenue Contribution)
        pivot_df = df.pivot_table(
            index='ABC Class', 
            columns='Entropy_Level', 
            values='FY Revenue', 
            aggfunc='sum'
        ).fillna(0.0)
        
        abc_order = ['A', 'B', 'C']
        entropy_order = ['Low', 'Medium', 'High', 'Very High']
        
        # Re-index to ensure correct order and size
        pivot_df = pivot_df.reindex(index=abc_order, columns=entropy_order, fill_value=0.0)
        
        heatmap_data = {
            "x": entropy_order,
            "y": abc_order,
            "z": pivot_df.values.tolist()
        }
        
        return jsonify({
            "status": "success",
            "skus": skus_data,
            "kpis": kpis,
            "heatmap": heatmap_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": f"Analysis failed: {str(e)}"
        }), 500

@app.route('/api/sku_detail')
def sku_detail():
    try:
        sku = request.args.get('sku')
        if not sku:
            return jsonify({"status": "error", "message": "SKU parameter is required"}), 400
            
        sku_val = str(sku)
        segment = request.args.get('segment', 'All')
        
        conn = get_db_connection()
        
        query = r"""
        SELECT 
            "Year" AS fy,
            "Month" AS month,
            SUM("SalesQty") AS qty,
            SUM(COALESCE("Revenue", "Taxable Value")) AS revenue
        FROM revenue
        WHERE "SalesQty" > 0 AND "Module WP"::text = %s
          {segment_filter}
        GROUP BY "Year", "Month"
        ORDER BY "Year", "Month";
        """
        
        params = [sku_val]
        
        segment_filter = ""
        if segment != 'All':
            segment_filter = 'AND "Segment" = %s'
            params.append(segment)
            
        full_query = query.format(segment_filter=segment_filter)
        
        df = pd.read_sql_query(full_query, conn, params=params)
        conn.close()
        
        months_fy = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
        months_names = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        
        result = {}
        if not df.empty:
            years = df['fy'].unique()
            for yr in years:
                df_yr = df[df['fy'] == yr].set_index('month')
                qty_series = []
                rev_series = []
                for m in months_fy:
                    if m in df_yr.index:
                        qty_series.append(float(df_yr.loc[m, 'qty']))
                        rev_series.append(float(df_yr.loc[m, 'revenue']))
                    else:
                        qty_series.append(0.0)
                        rev_series.append(0.0)
                        
                result[yr] = {
                    "months": months_names,
                    "qty": qty_series,
                    "revenue": rev_series
                }
                
        return jsonify({
            "status": "success",
            "sku": sku_val,
            "yoy_data": result
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to retrieve SKU details: {str(e)}"
        }), 500

@app.route('/api/yoy_comparison')
def yoy_comparison():
    """Endpoint to run comparison metrics of all SKUs across two years."""
    try:
        fy1 = request.args.get('fy1', '2025-26')
        fy2 = request.args.get('fy2', '2026-27')
        segment = request.args.get('segment', 'All')
        min_months = int(request.args.get('min_months', 3))
        min_qty = float(request.args.get('min_qty', 0))
        winsorize_pct = int(request.args.get('winsorize_pct', 99))
        abc_a = float(request.args.get('abc_a', 0.80))
        abc_b = float(request.args.get('abc_b', 0.15))
        adi_cutoff = float(request.args.get('adi_cutoff', 1.32))
        cv2_cutoff = float(request.args.get('cv2_cutoff', 0.49))
        
        df1, _ = run_entropy_analysis(
            fy=fy1, segment=segment,
            min_months=min_months, min_qty=min_qty, winsorize_pct=winsorize_pct,
            abc_a=abc_a, abc_b=abc_b, adi_cutoff=adi_cutoff, cv2_cutoff=cv2_cutoff
        )
        df2, _ = run_entropy_analysis(
            fy=fy2, segment=segment,
            min_months=min_months, min_qty=min_qty, winsorize_pct=winsorize_pct,
            abc_a=abc_a, abc_b=abc_b, adi_cutoff=adi_cutoff, cv2_cutoff=cv2_cutoff
        )
        
        if df1.empty or df2.empty:
            return jsonify({
                "status": "success",
                "comparison": []
            })
            
        # Merge years
        merged = pd.merge(df1, df2, on='Module WP', suffixes=('_yr1', '_yr2'))
        
        comp_records = []
        for idx, row in merged.iterrows():
            q1, q2 = row['FY Qty_yr1'], row['FY Qty_yr2']
            r1, r2 = row['FY Revenue_yr1'], row['FY Revenue_yr2']
            
            qty_diff_pct = ((q2 - q1) / q1 * 100.0) if q1 > 0 else 0.0
            rev_diff_pct = ((r2 - r1) / r1 * 100.0) if r1 > 0 else 0.0
            
            comp_records.append({
                "Module WP": row['Module WP'],
                "Mat Desc": row['Mat Desc_yr1'],
                "Qty_Yr1": q1,
                "Qty_Yr2": q2,
                "Qty_Diff_Pct": qty_diff_pct,
                "Revenue_Yr1": r1,
                "Revenue_Yr2": r2,
                "Revenue_Diff_Pct": rev_diff_pct,
                "Entropy_Yr1": row['H_norm_yr1'],
                "Entropy_Yr2": row['H_norm_yr2']
            })
            
        return jsonify({
            "status": "success",
            "comparison": comp_records
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"YoY calculation failed: {str(e)}"
        }), 500

@app.route('/api/export')
def export():
    try:
        fy = request.args.get('fy', '2025-26')
        segment = request.args.get('segment', 'All')
        min_months = int(request.args.get('min_months', 3))
        min_qty = float(request.args.get('min_qty', 0))
        winsorize_pct = int(request.args.get('winsorize_pct', 99))
        abc_a = float(request.args.get('abc_a', 0.80))
        abc_b = float(request.args.get('abc_b', 0.15))
        adi_cutoff = float(request.args.get('adi_cutoff', 1.32))
        cv2_cutoff = float(request.args.get('cv2_cutoff', 0.49))
        
        df, kpis = run_entropy_analysis(
            fy=fy, segment=segment,
            min_months=min_months, min_qty=min_qty, 
            winsorize_pct=winsorize_pct, abc_a=abc_a, abc_b=abc_b,
            adi_cutoff=adi_cutoff, cv2_cutoff=cv2_cutoff
        )
        
        if df.empty:
            wb = Workbook()
            ws = wb.active
            ws.title = "No Data Found"
            ws['A1'] = "No records matched the selected filters."
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, download_name="sku_entropy_empty.xlsx", as_attachment=True)
            
        wb = Workbook()
        
        # 1. Summary Sheet
        ws_sum = wb.active
        ws_sum.title = "KPI Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws_sum.merge_cells('A1:D1')
        ws_sum['A1'] = "SKU Temporal Entropy & Demand Analytics"
        ws_sum['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        ws_sum['A1'].fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_sum.row_dimensions[1].height = 40
        
        # Subtitle
        ws_sum['A2'] = f"FY: {fy} | Segment: {segment}"
        ws_sum['A2'].font = Font(name='Segoe UI', size=11, italic=True)
        ws_sum.merge_cells('A2:D2')
        ws_sum.row_dimensions[2].height = 20
        
        # Headers Style
        hdr_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        hdr_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        ws_sum['A4'] = "Key Performance Indicators"
        ws_sum['A4'].font = Font(name='Segoe UI', size=13, bold=True, color='1E3A8A')
        
        kpis_rows = [
            ("Total Active SKUs", kpis["total_skus"], "NOS"),
            ("Total FY Qty Sold", kpis["total_qty"], "#,##0.00"),
            ("Total FY Revenue", kpis["total_revenue"], "₹#,##0.00"),
            ("Avg Normalized Entropy", kpis["avg_entropy"], "0.00%")
        ]
        
        for idx, (label, val, fmt) in enumerate(kpis_rows, start=5):
            ws_sum.cell(row=idx, column=1, value=label).font = Font(name='Segoe UI', size=11, bold=True)
            cell = ws_sum.cell(row=idx, column=2, value=val)
            cell.font = Font(name='Segoe UI', size=11)
            cell.number_format = fmt
            ws_sum.cell(row=idx, column=1).border = thin_border
            cell.border = thin_border
            
        # SBC Counts Table
        ws_sum['A11'] = "SBC Demand Patterns"
        ws_sum['A11'].font = Font(name='Segoe UI', size=13, bold=True, color='1E3A8A')
        
        ws_sum.cell(row=12, column=1, value="Demand Pattern").fill = hdr_fill
        ws_sum.cell(row=12, column=1).font = hdr_font
        ws_sum.cell(row=12, column=2, value="SKU Count").fill = hdr_fill
        ws_sum.cell(row=12, column=2).font = hdr_font
        ws_sum.cell(row=12, column=3, value="Total Revenue").fill = hdr_fill
        ws_sum.cell(row=12, column=3).font = hdr_font
        
        patterns = ['Smooth', 'Intermittent', 'Erratic', 'Lumpy']
        for idx, pat in enumerate(patterns, start=13):
            c_cnt = kpis["sbc_counts"].get(pat, 0)
            c_rev = kpis["sbc_revenue"].get(pat, 0.0)
            
            ws_sum.cell(row=idx, column=1, value=pat).font = Font(name='Segoe UI', size=11)
            cell_cnt = ws_sum.cell(row=idx, column=2, value=c_cnt)
            cell_cnt.font = Font(name='Segoe UI', size=11)
            cell_cnt.number_format = '#,##0'
            cell_rev = ws_sum.cell(row=idx, column=3, value=c_rev)
            cell_rev.font = Font(name='Segoe UI', size=11)
            cell_rev.number_format = '₹#,##0.00'
            
            for col in range(1, 4):
                ws_sum.cell(row=idx, column=col).border = thin_border
                
        # ABC Table
        ws_sum['A19'] = "ABC Pareto Class Distribution"
        ws_sum['A19'].font = Font(name='Segoe UI', size=13, bold=True, color='1E3A8A')
        ws_sum.cell(row=20, column=1, value="ABC Class").fill = hdr_fill
        ws_sum.cell(row=20, column=1).font = hdr_font
        ws_sum.cell(row=20, column=2, value="SKU Count").fill = hdr_fill
        ws_sum.cell(row=20, column=2).font = hdr_font
        ws_sum.cell(row=20, column=3, value="Total Revenue").fill = hdr_fill
        ws_sum.cell(row=20, column=3).font = hdr_font
        
        classes = ['A', 'B', 'C']
        for idx, cls in enumerate(classes, start=21):
            c_cnt = kpis["abc_counts"].get(cls, 0)
            c_rev = kpis["abc_revenue"].get(cls, 0.0)
            
            ws_sum.cell(row=idx, column=1, value=f"Class {cls}").font = Font(name='Segoe UI', size=11)
            cell_cnt = ws_sum.cell(row=idx, column=2, value=c_cnt)
            cell_cnt.font = Font(name='Segoe UI', size=11)
            cell_cnt.number_format = '#,##0'
            cell_rev = ws_sum.cell(row=idx, column=3, value=c_rev)
            cell_rev.font = Font(name='Segoe UI', size=11)
            cell_rev.number_format = '₹#,##0.00'
            
            for col in range(1, 4):
                ws_sum.cell(row=idx, column=col).border = thin_border
                
        ws_sum.column_dimensions['A'].width = 25
        ws_sum.column_dimensions['B'].width = 18
        ws_sum.column_dimensions['C'].width = 22
        
        # 2. Details Sheet
        ws_det = wb.create_sheet(title="SKU Analysis Metrics")
        ws_det.views.sheetView[0].showGridLines = True
        
        columns_to_export = [
            'Module WP', 'Mat Desc', 'FY Qty', 'FY Revenue', 'ABC Class', 
            'Entropy', 'H_norm', 'NonZeroMonths', 'Entropy_Level', 'Overall_Segment', 'Actionable Recommendation'
        ]
        
        for col_idx, col_name in enumerate(columns_to_export, start=1):
            cell = ws_det.cell(row=1, column=col_idx, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws_det.row_dimensions[1].height = 28
        
        format_map = {
            'Module WP': '0',
            'FY Qty': '#,##0.00',
            'FY Revenue': '₹#,##0.00',
            'Entropy': '0.0000',
            'H_norm': '0.00%',
            'NonZeroMonths': '0'
        }
        
        zebra_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        
        for row_idx, row in enumerate(df[columns_to_export].itertuples(index=False), start=2):
            is_zebra = (row_idx % 2 == 0)
            ws_det.row_dimensions[row_idx].height = 20
            
            for col_idx, val in enumerate(row, start=1):
                col_name = columns_to_export[col_idx-1]
                cell = ws_det.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = thin_border
                
                if is_zebra:
                    cell.fill = zebra_fill
                    
                if col_name in format_map:
                    cell.number_format = format_map[col_name]
                    
                if col_name in ['Mat Desc', 'Actionable Recommendation', 'Overall_Segment', 'Entropy_Level', 'ABC Class']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
        for col in ws_det.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_det.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"sku_entropy_analysis_{fy}_{plant}_{segment}.xlsx".replace(" ", "_")
        return send_file(output, download_name=filename, as_attachment=True)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Export failed: {str(e)}"}), 500

@app.route('/api/sql')
def get_sql():
    sql_queries = {
        "full_entropy_analysis": r"""
-- Complete production-ready PostgreSQL query with Revenue Shannon Entropy and ABC Pareto classification
WITH base_sales AS (
    SELECT 
        "Module WP"::text AS sku,
        "Year" AS fy,
        "Month" AS month,
        "SalesQty" AS qty,
        COALESCE("Revenue", "Taxable Value") AS revenue_val,
        "Mat Desc" AS mat_desc
    FROM revenue
    WHERE "SalesQty" > 0 AND "Module WP" IS NOT NULL
      AND "Year" = '2025-26'
),
months AS (
    SELECT unnest(ARRAY[4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]) AS month
),
sku_info AS (
    SELECT sku, fy, MAX(mat_desc) AS mat_desc
    FROM base_sales
    GROUP BY sku, fy
),
skeleton AS (
    SELECT s.sku, s.mat_desc, s.fy, m.month
    FROM sku_info s
    CROSS JOIN months m
),
monthly_sales AS (
    SELECT sku, fy, month, SUM(qty) AS qty_sum, SUM(revenue_val) AS rev_sum
    FROM base_sales
    GROUP BY sku, fy, month
),
filled_sales AS (
    SELECT 
        s.sku, s.mat_desc, s.fy, s.month,
        COALESCE(m.qty_sum, 0) AS monthly_qty,
        COALESCE(m.rev_sum, 0) AS monthly_revenue
    FROM skeleton s
    LEFT JOIN monthly_sales m ON s.sku = m.sku AND s.fy = m.fy AND s.month = m.month
),
sku_totals AS (
    SELECT 
        sku, mat_desc, fy,
        SUM(monthly_qty) AS total_qty,
        SUM(monthly_revenue) AS total_revenue,
        COUNT(CASE WHEN monthly_revenue > 0 THEN 1 END) AS nonzero_months
    FROM filled_sales
    GROUP BY sku, mat_desc, fy
),
monthly_probabilities AS (
    SELECT 
        f.sku, f.mat_desc, f.fy, f.month, f.monthly_qty, f.monthly_revenue,
        t.total_qty, t.total_revenue,
        CASE WHEN t.total_revenue > 0 AND f.monthly_revenue > 0 THEN f.monthly_revenue / t.total_revenue ELSE 0 END AS p_rev
    FROM filled_sales f
    JOIN sku_totals t ON f.sku = t.sku AND f.mat_desc = t.mat_desc AND f.fy = t.fy
),
sku_entropy AS (
    SELECT 
        sku, mat_desc, fy,
        -SUM(CASE WHEN p_rev > 0 THEN p_rev * (LN(p_rev) / LN(2)) ELSE 0 END) AS entropy_rev
    FROM monthly_probabilities
    GROUP BY sku, mat_desc, fy
),
sku_metrics AS (
    SELECT 
        t.sku, t.mat_desc, t.fy, t.total_qty, t.total_revenue, t.nonzero_months,
        e.entropy_rev,
        e.entropy_rev / 3.58496250072 AS h_norm,
        SUM(t.total_revenue) OVER (PARTITION BY t.fy) AS fy_total_revenue
    FROM sku_totals t
    JOIN sku_entropy e ON t.sku = e.sku AND t.mat_desc = e.mat_desc AND t.fy = e.fy
),
sku_ranked AS (
    SELECT *,
        SUM(total_revenue) OVER (PARTITION BY fy ORDER BY total_revenue DESC ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) AS running_revenue
    FROM sku_metrics
),
sku_abc AS (
    SELECT *,
        CASE WHEN fy_total_revenue > 0 THEN running_revenue / fy_total_revenue ELSE 0 END AS cum_revenue_pct
    FROM sku_ranked
)
SELECT 
    sku AS "Module WP",
    mat_desc AS "Mat Desc",
    fy AS "FY",
    total_qty AS "FY Qty",
    total_revenue AS "FY Revenue",
    CASE 
        WHEN cum_revenue_pct <= 0.80 THEN 'A'
        WHEN cum_revenue_pct - (total_revenue / NULLIF(fy_total_revenue, 0)) <= 0.80 THEN 'A'
        WHEN cum_revenue_pct - (total_revenue / NULLIF(fy_total_revenue, 0)) <= 0.95 THEN 'B'
        ELSE 'C'
    END AS "ABC Class",
    entropy_rev AS "Entropy",
    h_norm AS "H_norm",
    nonzero_months AS "NonZeroMonths",
    CASE 
        WHEN h_norm < 0.30 THEN 'Low'
        WHEN h_norm >= 0.30 AND h_norm < 0.60 THEN 'Medium'
        WHEN h_norm >= 0.60 AND h_norm < 0.80 THEN 'High'
        ELSE 'Very High'
    END AS "Entropy_Level"
FROM sku_abc
ORDER BY fy, total_revenue DESC;
        """
    }
    return jsonify(sql_queries)

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
